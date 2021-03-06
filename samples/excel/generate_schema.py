#!/usr/bin/env python
# -*- encoding: utf-8 -*-
"""
Topic: 通过一个schema.sql来生成excel表格的数据库设计文档
Desc :
"""
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.styles import colors, borders, fills
import re
from copy import copy


def load_schema(filename):
    """先加载schema.sql文件来获取所有建表语句"""
    result = []
    pat = re.compile(r'.* DEFAULT (\S+) .*')
    with open(filename, encoding='utf-8') as sqlfile:
        each_table = []  # 每张表定义
        temp_comment = ''
        for line in sqlfile:
            if not line.strip() or line.strip().startswith("#"):
                continue
            line = line.replace("`", "")
            if line.startswith('--'):
                temp_comment = line.split('--')[1].strip()
            elif 'DROP TABLE' in line:
                each_table.insert(0, temp_comment)
                each_table.insert(1, line.strip().split()[-1][:-1])
            elif ' COMMENT ' in line and 'ENGINE=' not in line:
                col_arr = line.split()
                col_name = col_arr[0]
                col_type = col_arr[1]
                if 'PRIMARY KEY' in line or 'NOT NULL' in line:
                    col_null = 'NOT NULL'
                else:
                    col_null = ''
                col_remark = line.split(' COMMENT ')
                cr = col_remark[-1].strip().replace("'", "")
                defaultmatch = pat.match(line)
                default = defaultmatch.group(1) if defaultmatch else ''
                each_table.append((col_name, col_type, col_null,
                                   default, cr[:-1] if cr.endswith(',') else cr))
            elif 'ENGINE=' in line:
                # 单个表定义结束
                result.append(list(each_table))
                each_table.clear()
    return result


def write_dest(xlsx_name, schema_name):
    border = Border(
        left=Side(border_style=borders.BORDER_THIN, color='FF000000'),
        right=Side(border_style=borders.BORDER_THIN, color='FF000000'),
        top=Side(border_style=borders.BORDER_THIN, color='FF000000'),
        bottom=Side(border_style=borders.BORDER_THIN, color='FF000000')
    )
    alignment = Alignment(horizontal='justify', vertical='bottom',
                          text_rotation=0, wrap_text=False,
                          shrink_to_fit=True, indent=0)
    fill = PatternFill(fill_type=None, start_color='FFFFFFFF')
    # 基本的样式
    basic_style = NamedStyle(name="basic_style", font=Font(name='Microsoft YaHei')
                             , border=border, alignment=alignment, fill=fill)
    title_style = copy(basic_style)
    title_style.name = 'title_style'
    title_style.font = Font(name='Microsoft YaHei', b=True, size=20, color='00215757')
    title_style.alignment = Alignment(horizontal='center', vertical='bottom',
                                      text_rotation=0, wrap_text=False,
                                      shrink_to_fit=True, indent=0)
    title_style.fill = PatternFill(fill_type=fills.FILL_SOLID, start_color='00B2CBED')
    header_style = copy(basic_style)
    header_style.name='header_style'
    header_style.font = Font(name='Microsoft YaHei', b=True, size=15, color='00215757')
    header_style.fill = PatternFill(fill_type=fills.FILL_SOLID, start_color='00BAA87F')
    common_style = copy(basic_style)
    common_style.name = 'common_style'
    link_style = copy(basic_style)
    link_style.name = 'link_style'
    link_style.font = Font(name='Microsoft YaHei', color=colors.BLUE, underline='single')
    table_data = load_schema(schema_name)
    wb = Workbook()
    wb.add_named_style(basic_style)
    wb.add_named_style(title_style)
    wb.add_named_style(header_style)
    wb.add_named_style(common_style)

    bstyle = NamedStyle(name='bstyle', border=Border(
            bottom=Side(border_style=borders.BORDER_THIN, color='FF000000')))
    wb.add_named_style(bstyle)

    wb.active.title = "首页列表"

    for table in table_data:
        ws = wb.create_sheet(title=table[0])
        ws.merge_cells('E3:I3')  # 合并单元格
        ws['E3'].style = title_style
        ws['F2'].style = bstyle
        ws['G2'].style = bstyle
        ws['H2'].style = bstyle
        ws['I2'].style = bstyle
        ws['J3'].style = bstyle
        ws['E3'] = table[0]
        ws['E4'].style = header_style
        ws['E4'] = '列名'
        ws['F4'].style = header_style
        ws['F4'] = '类型'
        ws['G4'].style = header_style
        ws['G4'] = '空值约束'
        ws['H4'].style = header_style
        ws['H4'] = '默认值'
        ws['I4'].style = header_style
        ws['I4'] = '备注'
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 25
        ws.column_dimensions['I'].width = 40
        for idx, each_column in enumerate(table[2:]):
            ws['E{}'.format(idx + 5)].style = common_style
            ws['E{}'.format(idx + 5)] = each_column[0]
            ws['F{}'.format(idx + 5)].style = common_style
            ws['F{}'.format(idx + 5)] = each_column[1]
            ws['G{}'.format(idx + 5)].style = common_style
            ws['G{}'.format(idx + 5)] = each_column[2]
            ws['H{}'.format(idx + 5)].style = common_style
            ws['H{}'.format(idx + 5)] = each_column[3]
            ws['I{}'.format(idx + 5)].style = common_style
            ws['I{}'.format(idx + 5)] = each_column[4]
    ws = wb['首页列表']
    ws.merge_cells('D3:F3')
    ws['D3'].style = title_style
    ws['E2'].style = bstyle
    ws['F2'].style = bstyle
    ws['G3'].style = bstyle
    ws['D3'] = 'MySQL数据库系统表'
    ws['D4'].style = header_style
    ws['D4'] = '编号'
    ws['E4'].style = header_style
    ws['E4'] = '表名'
    ws['F4'].style = header_style
    ws['F4'] = '详情链接'
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 45
    for inx, val in enumerate(table_data):
        ws['D{}'.format(inx + 5)].style = common_style
        ws['D{}'.format(inx + 5)] = inx + 1
        ws['E{}'.format(inx + 5)].style = common_style
        ws['E{}'.format(inx + 5)] = val[1]
        linkcell = ws['F{}'.format(inx + 5)]
        linkcell.style = link_style
        linkcell.value = val[0]
        linkcell.hyperlink = '#{0}!{1}'.format(val[0], 'E3')
    wb.save(filename=xlsx_name)


if __name__ == '__main__':
    # write_xlsx()
    # write_only()
    import sys

    dest_file = r'E:\work\MySQL数据库设计.xlsx'
    schema_file = r'E:\projects\clouds-epay-mapper\sql\schema.sql'
    write_dest(dest_file, schema_file)
    # write_dest(sys.argv[1], sys.argv[2])
    pass
